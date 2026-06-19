"""
web_report.py — Production Analytics Web Dashboard
Runs Flask on localhost:5001 and auto-opens browser.
"""
import sys
import os
import io
import threading
import webbrowser
from datetime import datetime

if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    WORK_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    WORK_DIR = BASE_DIR

sys.path.insert(0, BASE_DIR)

from flask import Flask, render_template, request, jsonify, send_file, Response
from shared_config import DB, make_logger, APP_TITLE, APP_VERSION

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

PORT = 5001
log = make_logger("web_report")
db  = DB(log)

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static'),
)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

DETAILED_COLS = [
    "Log ID", "Production ID", "Timestamp", "Weight (g)", "Status",
    "Metal Status", "Lot No", "Product", "Shift", "Batch No",
    "Net Weight", "Under Limit", "Over Limit", "Machine",
]
SUMMARY_COLS = [
    "Production ID", "Machine", "Lot No", "Product", "Shift", "Batch No",
    "First Reading", "Last Reading", "Total Readings", "Pass Count",
    "Under Fail", "Over Fail", "Metal Fail", "Pass Rate %",
    "Min Weight", "Max Weight", "Avg Weight",
]

# ── helpers ──────────────────────────────────────────────────────────────────
def _fmt(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return '' if v is None else v

def _rows_json(rows):
    return [[_fmt(v) for v in row] for row in rows]

def get_machines():
    try:
        r = db.call_sp("sp_GetAllMachines", fetch=True)
        return [row[0] for row in r] if r else []
    except Exception:
        return []

def run_report(machine, start, end, lot, batch, rtype):
    return db.call_sp("sp_GetProductionReport", [
        machine or None, start or None, end or None,
        lot or None, batch or None, rtype,
    ], fetch=True) or []

# ── analytics api ────────────────────────────────────────────────────────────
@app.route('/api/analytics')
def api_analytics():
    machine = request.args.get('machine', '')
    start   = request.args.get('start', datetime.now().strftime('%Y-%m-01'))
    end     = request.args.get('end',   datetime.now().strftime('%Y-%m-%d'))
    try:
        rows = run_report(machine, start, end, '', '', 'monthly')

        # ── FPY trend by date ──────────────────────────────────────────────────
        fpy_map = {}
        for r in rows:
            d = r[6]
            if not d: continue
            day = d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)[:10]
            e = fpy_map.setdefault(day, {'bags': 0, 'pass': 0})
            e['bags'] += int(r[8] or 0)
            e['pass'] += int(r[9] or 0)
        fpy_trend = [{'date': d[:5] if len(d) > 5 else d,
                      'fpy': round(v['pass'] / v['bags'] * 100, 1) if v['bags'] else 0,
                      'bags': v['bags']}
                     for d, v in sorted(fpy_map.items())]

        # ── defect pareto ──────────────────────────────────────────────────────
        total_under = sum(int(r[10] or 0) for r in rows)
        total_over  = sum(int(r[11] or 0) for r in rows)
        total_metal = sum(int(r[12] or 0) for r in rows)

        # ── shift breakdown ────────────────────────────────────────────────────
        shift_map = {}
        for r in rows:
            s = (r[4] or 'Unknown').strip() or 'Unknown'
            e = shift_map.setdefault(s, {'bags': 0, 'pass': 0, 'under': 0, 'over': 0, 'metal': 0})
            e['bags']  += int(r[8]  or 0)
            e['pass']  += int(r[9]  or 0)
            e['under'] += int(r[10] or 0)
            e['over']  += int(r[11] or 0)
            e['metal'] += int(r[12] or 0)
        shifts = sorted([
            {'shift': s, 'bags': v['bags'],
             'rate': round(v['pass'] / v['bags'] * 100, 1) if v['bags'] else 0,
             'defects': v['under'] + v['over'] + v['metal']}
            for s, v in shift_map.items()
        ], key=lambda x: x['bags'], reverse=True)

        # ── machine comparison ─────────────────────────────────────────────────
        mach_map = {}
        for r in rows:
            m = r[1] or 'Unknown'
            e = mach_map.setdefault(m, {'bags': 0, 'pass': 0, 'under': 0, 'over': 0, 'metal': 0})
            e['bags']  += int(r[8]  or 0)
            e['pass']  += int(r[9]  or 0)
            e['under'] += int(r[10] or 0)
            e['over']  += int(r[11] or 0)
            e['metal'] += int(r[12] or 0)
        machines_cmp = [{'machine': m, **v,
                          'rate': round(v['pass'] / v['bags'] * 100, 1) if v['bags'] else 0}
                        for m, v in mach_map.items()]

        # ── product breakdown ──────────────────────────────────────────────────
        prod_map = {}
        for r in rows:
            p = (r[3] or 'Unknown').strip() or 'Unknown'
            e = prod_map.setdefault(p, {'bags': 0, 'pass': 0})
            e['bags'] += int(r[8] or 0)
            e['pass'] += int(r[9] or 0)
        products = sorted([
            {'product': p,
             'bags': v['bags'],
             'rate': round(v['pass'] / v['bags'] * 100, 1) if v['bags'] else 0}
            for p, v in prod_map.items()
        ], key=lambda x: x['bags'], reverse=True)[:8]

        # ── hourly production distribution ─────────────────────────────────────
        hourly = {}
        for r in rows:
            d = r[6]
            if d and hasattr(d, 'hour'):
                h = d.hour
                hourly[h] = hourly.get(h, 0) + int(r[8] or 0)
        hourly_data = [{'hour': f"{h:02d}:00", 'bags': hourly.get(h, 0)} for h in range(24)]

        # ── candlestick OHLC per session (trading-style weight chart) ─────────────
        # high=max_weight, low=min_weight
        # body (open/close) = Q1/Q3 approximation of weight distribution
        # color = FPY quality: green ≥95%, orange 80-95%, red <80%
        candles = []
        for r in sorted(rows, key=lambda x: x[6] if x[6] else datetime.min):
            if r[16] is None or r[14] is None or r[15] is None or r[6] is None:
                continue
            ts   = r[6]
            mn   = float(r[14])
            mx   = float(r[15])
            avg  = float(r[16])
            span = mx - mn
            rate = round(float(r[13] or 0), 1)
            # body represents middle 50% of weight distribution
            op = round(mn + span * 0.25, 1)
            cl = round(mn + span * 0.75, 1)
            if op == cl:                  # prevent zero-height body
                op = round(avg - 0.5, 1)
                cl = round(avg + 0.5, 1)
            color = '#30D158' if rate >= 95 else '#FF9F0A' if rate >= 80 else '#FF453A'
            candles.append({
                'time':   int(ts.timestamp()) if hasattr(ts, 'timestamp') else 0,
                'open':   op,
                'high':   mx,
                'low':    mn,
                'close':  cl,
                'color':  color,
                'volume': int(r[8] or 0),
                'rate':   rate,
                'lot':    r[2] or '',
                'avg':    round(avg, 1),
            })

        # ── top 5 sessions by volume ───────────────────────────────────────────
        top = sorted(rows, key=lambda r: int(r[8] or 0), reverse=True)[:5]
        top_sessions = [{'lot': r[2], 'product': r[3], 'shift': r[4] or '—',
                         'machine': r[1] or '—', 'bags': int(r[8] or 0),
                         'rate': round(float(r[13] or 0), 1)}
                        for r in top]

        # ── throughput (bags/hour per session) avg ─────────────────────────────
        throughputs = []
        for r in rows:
            t1, t2 = r[6], r[7]
            bags = int(r[8] or 0)
            if t1 and t2 and bags > 0 and hasattr(t1, 'hour') and hasattr(t2, 'hour'):
                hrs = (t2 - t1).total_seconds() / 3600
                if hrs > 0.05:
                    throughputs.append(bags / hrs)
        avg_throughput = round(sum(throughputs) / len(throughputs), 1) if throughputs else 0

        return jsonify({
            'fpy_trend': fpy_trend,
            'defect_pareto': {'labels': ['Under Weight', 'Over Weight', 'Metal Fail'],
                              'data': [total_under, total_over, total_metal]},
            'shifts': shifts,
            'machines': machines_cmp,
            'products': products,
            'hourly': hourly_data,
            'candles': candles,
            'top_sessions': top_sessions,
            'avg_throughput': avg_throughput,
        })
    except Exception as e:
        log.error(f"Analytics API: {e}")
        return jsonify({'error': str(e)}), 500

# ── pages ────────────────────────────────────────────────────────────────────
@app.route('/')
def dashboard():
    return render_template('dashboard.html', machines=get_machines(),
                           title=APP_TITLE, version=APP_VERSION, active='dashboard')

@app.route('/reports')
def reports_page():
    return render_template('reports.html', machines=get_machines(),
                           title=APP_TITLE, version=APP_VERSION, active='reports',
                           today=datetime.now().strftime('%Y-%m-%d'),
                           month_start=datetime.now().strftime('%Y-%m-01'))

# ── api ──────────────────────────────────────────────────────────────────────
@app.route('/api/dashboard')
def api_dashboard():
    machine = request.args.get('machine', '')
    start   = request.args.get('start', datetime.now().strftime('%Y-%m-01'))
    end     = request.args.get('end',   datetime.now().strftime('%Y-%m-%d'))
    try:
        rows        = run_report(machine, start, end, '', '', 'monthly')
        total_bags  = sum(int(r[8]  or 0) for r in rows)
        total_pass  = sum(int(r[9]  or 0) for r in rows)
        total_under = sum(int(r[10] or 0) for r in rows)
        total_over  = sum(int(r[11] or 0) for r in rows)
        total_metal = sum(int(r[12] or 0) for r in rows)
        pass_rate   = round(total_pass / total_bags * 100, 1) if total_bags else 0

        daily = {}
        for r in rows:
            d = r[6]
            if d:
                day = d.strftime('%b %d') if hasattr(d, 'strftime') else str(d)[:10]
                daily[day] = daily.get(day, 0) + int(r[8] or 0)
        daily_sorted = sorted(daily.items(), key=lambda x: x[0])

        recent = [{
            'prod_id': r[0], 'machine': r[1] or '', 'lot': r[2] or '',
            'product': r[3] or '', 'shift': r[4] or '',
            'total': int(r[8] or 0), 'pass': int(r[9] or 0),
            'under': int(r[10] or 0), 'over': int(r[11] or 0),
            'metal': int(r[12] or 0),
            'rate': round(float(r[13] or 0), 1),
            'first': _fmt(r[6])[:16], 'last': _fmt(r[7])[:16],
        } for r in rows[:20]]

        return jsonify({
            'kpis': {
                'sessions': len(rows), 'bags': total_bags, 'pass_rate': pass_rate,
                'under': total_under, 'over': total_over, 'metal': total_metal,
            },
            'donut': {
                'labels': ['Pass', 'Under Weight', 'Over Weight'],
                'data': [total_pass, total_under, total_over],
            },
            'daily': {
                'labels': [d[0] for d in daily_sorted],
                'data':   [d[1] for d in daily_sorted],
            },
            'recent': recent,
        })
    except Exception as e:
        log.error(f"Dashboard API: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/report')
def api_report():
    machine = request.args.get('machine', '')
    start   = request.args.get('start', '')
    end     = request.args.get('end', '')
    lot     = request.args.get('lot', '')
    batch   = request.args.get('batch', '')
    rtype   = request.args.get('type', 'detailed')
    try:
        rows = run_report(machine, start, end, lot, batch, rtype)
        return jsonify({'rows': _rows_json(rows), 'type': rtype, 'count': len(rows)})
    except Exception as e:
        log.error(f"Report API: {e}")
        return jsonify({'error': str(e)}), 500

# ── excel export ─────────────────────────────────────────────────────────────
@app.route('/export/excel')
def export_excel():
    if not HAS_EXCEL:
        return Response("openpyxl not installed", 500)
    machine = request.args.get('machine', '')
    start   = request.args.get('start', '')
    end     = request.args.get('end', '')
    lot     = request.args.get('lot', '')
    batch   = request.args.get('batch', '')
    rtype   = request.args.get('type', 'detailed')
    try:
        rows = run_report(machine, start, end, lot, batch, rtype)
        buf  = _build_excel(rows, rtype)
        buf.seek(0)
        fname = f"report_{rtype}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        log.error(f"Excel export: {e}")
        return Response(f"Error: {e}", 500)

def _build_excel(rows, rtype):
    cols = list(DETAILED_COLS if rtype == 'detailed' else SUMMARY_COLS)
    if rows:
        nc = len(rows[0])
        if nc > len(cols): cols += [f"Col_{i}" for i in range(len(cols), nc)]
        else:              cols = cols[:nc]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Detailed' if rtype == 'detailed' else 'Summary'

    hdr_fill = PatternFill("solid", fgColor="7C5CFC")
    hdr_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)

    ws.cell(row=1, column=1,
            value=f"PanCen Production Report — {ws.title} — {datetime.now():%Y-%m-%d %H:%M}")
    ws.cell(row=1, column=1).font = title_font

    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    status_map = {0: 'Under', 1: 'Pass', 2: 'Over'}
    metal_map  = {0: 'Pass',  1: 'Fail'}

    for r_idx, row in enumerate(rows, 3):
        for c_idx, v in enumerate(row):
            if rtype == 'detailed':
                if c_idx == 4: v = status_map.get(v, v)
                elif c_idx == 5: v = metal_map.get(v, 'N/A') if v is not None else 'N/A'
            if hasattr(v, 'strftime'): v = v.strftime('%Y-%m-%d %H:%M:%S')
            elif v is None: v = ''
            ws.cell(row=r_idx, column=c_idx + 1, value=v)

    for col_cells in ws.columns:
        maxw = max((len(str(c.value or '')) for c in col_cells if c.row >= 2), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(maxw + 3, 45)

    if rtype != 'detailed' and rows:
        _excel_stats_sheet(wb, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf

def _excel_stats_sheet(wb, rows):
    ws = wb.create_sheet("Statistics")
    total_bags  = sum(int(r[8]  or 0) for r in rows)
    total_pass  = sum(int(r[9]  or 0) for r in rows)
    total_under = sum(int(r[10] or 0) for r in rows)
    total_over  = sum(int(r[11] or 0) for r in rows)
    total_metal = sum(int(r[12] or 0) for r in rows)
    pass_rate   = total_pass / total_bags * 100 if total_bags else 0
    min_w = min((r[14] for r in rows if r[14] is not None), default=0)
    max_w = max((r[15] for r in rows if r[15] is not None), default=0)
    avg_vals = [float(r[16]) for r in rows if r[16] is not None]
    avg_w = sum(avg_vals) / len(avg_vals) if avg_vals else 0
    data = [
        ("PRODUCTION STATISTICS SUMMARY", ""),
        ("Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("", ""),
        ("Total Sessions",    len(rows)),
        ("Total Readings",    total_bags),
        ("Total Pass",        total_pass),
        ("Overall Pass Rate", f"{pass_rate:.2f}%"),
        ("Total Under Fail",  total_under),
        ("Total Over Fail",   total_over),
        ("Total Metal Fail",  total_metal),
        ("", ""),
        ("Min Weight (g)",    min_w),
        ("Max Weight (g)",    max_w),
        ("Avg Weight (g)",    f"{avg_w:.1f}"),
    ]
    for r_idx, (metric, value) in enumerate(data, 1):
        ws.cell(row=r_idx, column=1, value=metric)
        ws.cell(row=r_idx, column=2, value=value)
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

# ── pdf export ───────────────────────────────────────────────────────────────
@app.route('/export/pdf')
def export_pdf():
    if not HAS_PDF:
        return Response("reportlab not installed", 500)
    machine = request.args.get('machine', '')
    start   = request.args.get('start', '')
    end     = request.args.get('end', '')
    lot     = request.args.get('lot', '')
    batch   = request.args.get('batch', '')
    rtype   = request.args.get('type', 'detailed')
    try:
        rows = run_report(machine, start, end, lot, batch, rtype)
        buf  = io.BytesIO()
        _build_pdf(buf, rows, rtype, machine or 'All', start, end)
        buf.seek(0)
        fname = f"report_{rtype}_{datetime.now():%Y%m%d_%H%M}.pdf"
        return send_file(buf, mimetype='application/pdf',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        log.error(f"PDF export: {e}")
        return Response(f"Error: {e}", 500)

def _build_pdf(buf, rows, rtype, machine_label, start, end):
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    elems  = []
    styles = getSampleStyleSheet()
    DARK, PURPLE = '#2d2d44', '#7c5cfc'

    hdr = Table([[
        Paragraph(f"PanCen Production — {'Detailed' if rtype=='detailed' else 'Summary'} Report",
                  ParagraphStyle('T', parent=styles['Title'], fontSize=14, textColor=colors.white)),
        Paragraph(f"Generated: {datetime.now():%Y-%m-%d %H:%M}",
                  ParagraphStyle('D', parent=styles['Normal'], fontSize=9,
                                 textColor=colors.HexColor('#cccccc'), alignment=2)),
    ]], colWidths=[440, 200])
    hdr.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor(DARK)),
        ('TOPPADDING',(0,0),(-1,0),10),('BOTTOMPADDING',(0,0),(-1,0),10),
        ('LEFTPADDING',(0,0),(-1,0),12),('RIGHTPADDING',(0,0),(-1,0),12),
        ('VALIGN',(0,0),(-1,0),'MIDDLE'),
    ]))
    elems += [hdr, Spacer(1, 5*mm)]
    elems.append(Paragraph(
        f"Machine: {machine_label}  |  Date: {start or 'All'} → {end or 'All'}  |  Records: {len(rows)}",
        ParagraphStyle('FI', parent=styles['Normal'], fontSize=8,
                       textColor=colors.HexColor('#888888'))))
    elems.append(Spacer(1, 4*mm))
    if rtype != 'detailed' and rows:
        elems += _pdf_stats_box(rows, DARK, PURPLE)
        elems.append(Spacer(1, 5*mm))
    elems += _pdf_data_table(rows, rtype, PURPLE)
    elems.append(Spacer(1, 6*mm))
    elems.append(Paragraph(
        f"PanCen Software v{APP_VERSION}  |  {datetime.now():%Y-%m-%d %H:%M:%S}",
        ParagraphStyle('F', parent=styles['Normal'], fontSize=7,
                       textColor=colors.HexColor('#666666'), alignment=1)))
    doc.build(elems)

def _pdf_stats_box(rows, DARK, PURPLE):
    bags  = sum(int(r[8]  or 0) for r in rows)
    pass_ = sum(int(r[9]  or 0) for r in rows)
    under = sum(int(r[10] or 0) for r in rows)
    over  = sum(int(r[11] or 0) for r in rows)
    metal = sum(int(r[12] or 0) for r in rows)
    rate  = pass_ / bags * 100 if bags else 0
    min_w = min((r[14] for r in rows if r[14] is not None), default=0)
    max_w = max((r[15] for r in rows if r[15] is not None), default=0)
    avgs  = [float(r[16]) for r in rows if r[16] is not None]
    avg_w = sum(avgs) / len(avgs) if avgs else 0
    t = Table([
        ["Sessions","Bags","Pass","Under","Over","Metal Fail","Pass Rate","Min (g)","Max (g)","Avg (g)"],
        [str(len(rows)),str(bags),str(pass_),str(under),str(over),
         str(metal),f"{rate:.1f}%",str(min_w),str(max_w),f"{avg_w:.1f}"],
    ], colWidths=[58]*10)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor(PURPLE)),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),7),
        ('BACKGROUND',(0,1),(-1,1),colors.HexColor(DARK)),
        ('TEXTCOLOR',(0,1),(-1,1),colors.white),
        ('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),('FONTSIZE',(0,1),(-1,1),9),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#3d3d5c')),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
    ]))
    return [t]

def _pdf_data_table(rows, rtype, PURPLE):
    elems = []
    if rtype == 'detailed':
        headers = ["#","Prod ID","Timestamp","Weight","Status","Metal","Lot No","Product","Machine"]
        idx     = [0,1,2,3,4,5,6,7,13]
        widths  = [35,42,95,50,40,38,60,70,58]
    else:
        headers = ["Prod ID","Machine","Lot","Product","Total","Pass","Under","Over","Metal","Rate%","Min","Max","Avg"]
        idx     = [0,1,2,3,8,9,10,11,12,13,14,15,16]
        widths  = [40,52,52,58,36,36,36,36,36,40,42,42,42]

    formatted = []
    for row in rows:
        r = []
        for j in idx:
            v = row[j] if j < len(row) else None
            if hasattr(v, 'strftime'):         r.append(v.strftime('%m-%d %H:%M'))
            elif j == 4 and rtype=='detailed': r.append('Pass' if v==1 else 'Under' if v==0 else 'Over')
            elif j == 5 and rtype=='detailed': r.append('Fail' if v==1 else 'OK' if v==0 else str(v or ''))
            elif j in (13,16) and rtype!='detailed':
                r.append(f"{float(v):.1f}" if v is not None else '')
            else: r.append('' if v is None else str(v))
        formatted.append(r)

    for pg in range(0, max(len(formatted), 1), 45):
        chunk = formatted[pg:pg+45]
        t = Table([headers]+chunk, colWidths=widths, repeatRows=1)
        cmds = [
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor(PURPLE)),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),7),
            ('FONTNAME',(0,1),(-1,-1),'Helvetica'),('FONTSIZE',(0,1),(-1,-1),6.5),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#3d3d5c')),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),
             [colors.HexColor('#f8f8ff'),colors.HexColor('#eeeef8')]),
        ]
        if rtype == 'detailed':
            for i, r in enumerate(chunk, 1):
                if len(r) > 4:
                    if   r[4]=='Under': cmds.append(('TEXTCOLOR',(4,i),(4,i),colors.HexColor('#ff8c00')))
                    elif r[4]=='Over':  cmds.append(('TEXTCOLOR',(4,i),(4,i),colors.HexColor('#ff3333')))
                    elif r[4]=='Pass':  cmds.append(('TEXTCOLOR',(4,i),(4,i),colors.HexColor('#22aa44')))
                if len(r) > 5 and r[5]=='Fail':
                    cmds.append(('TEXTCOLOR',(5,i),(5,i),colors.HexColor('#ff3333')))
        t.setStyle(TableStyle(cmds))
        elems.append(t)
        if pg + 45 < len(formatted): elems.append(PageBreak())
    return elems

# ── shutdown ─────────────────────────────────────────────────────────────────
@app.route('/shutdown', methods=['POST'])
def shutdown():
    threading.Timer(0.5, lambda: os._exit(0)).start()
    return jsonify({'ok': True})

# ── launch ───────────────────────────────────────────────────────────────────
def main():
    if not db.connect():
        log.warning("DB connection failed — app will show errors on data load")
    from shared_config import get_local_ip
    local_ip = get_local_ip()
    threading.Timer(1.2, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
    log.info(f"Web report server on http://{local_ip}:{PORT} (network) / http://localhost:{PORT} (local)")
    app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False)

if __name__ == '__main__':
    main()
